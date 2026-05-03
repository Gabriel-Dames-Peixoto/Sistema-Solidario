import tempfile
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


def build_dashboard(data: Dict, matches: List[Dict]) -> Dict:
    delivered = [item for item in data["items"] if item["status"] == "entregue"]
    available = [item for item in data["items"] if item["status"] == "disponivel"]
    beneficiaries = {request["beneficiary_id"] for request in data["requests"] if request["status"] in {"aberta", "atendida"}}
    return {
        "total_users": len(data["users"]),
        "total_items": len(data["items"]),
        "available_items": len(available),
        "delivered_items": len(delivered),
        "open_requests": len([request for request in data["requests"] if request["status"] == "aberta"]),
        "beneficiaries_served": len(beneficiaries),
        "total_matches": len(matches),
    }


def export_excel(output_path: Path, dashboard: Dict, items: List[Dict], requests: List[Dict], matches: List[Dict]) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tempfile.tempdir = str(output_path.parent.resolve())
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Resumo"
    summary.append(["Indicador", "Valor"])
    for key, value in dashboard.items():
        summary.append([key, value])

    items_sheet = workbook.create_sheet("Itens")
    items_sheet.append(["ID", "Titulo", "Categoria", "Quantidade", "Status", "Regiao"])
    for item in items:
        items_sheet.append([item["id"], item["title"], item["category"], item["quantity"], item["status"], item["region"]])

    requests_sheet = workbook.create_sheet("Solicitacoes")
    requests_sheet.append(["ID", "Categoria", "Quantidade", "Status", "Regiao"])
    for request in requests:
        requests_sheet.append([request["id"], request["category"], request["needed_quantity"], request["status"], request["region"]])

    match_sheet = workbook.create_sheet("Matches")
    match_sheet.append(["Item", "Solicitacao", "Categoria", "Quantidade", "Distancia", "Score"])
    for match in matches:
        match_sheet.append(
            [
                match["item_id"],
                match["request_id"],
                match["category"],
                match["allocated_quantity"],
                match["distance"],
                match["score"],
            ]
        )

    workbook.save(output_path)
    return output_path


def export_pdf(output_path: Path, dashboard: Dict, matches: List[Dict]) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    pdf = canvas.Canvas(str(output_path), pagesize=A4)
    width, height = A4
    line_y = height - 50

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(50, line_y, "Relatorio de Impacto - Sistema Solidario")
    line_y -= 30

    pdf.setFont("Helvetica", 11)
    for key, value in dashboard.items():
        pdf.drawString(50, line_y, f"{key}: {value}")
        line_y -= 18

    line_y -= 10
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(50, line_y, "Principais combinacoes sugeridas")
    line_y -= 20
    pdf.setFont("Helvetica", 10)
    for match in matches[:10]:
        pdf.drawString(
            50,
            line_y,
            f"Item {match['item_id']} -> Solicitacao {match['request_id']} | {match['category']} | qtd {match['allocated_quantity']} | dist {match['distance']}",
        )
        line_y -= 16
        if line_y < 60:
            pdf.showPage()
            line_y = height - 50
            pdf.setFont("Helvetica", 10)

    pdf.save()
    return output_path
